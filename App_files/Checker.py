# from PyQt6.QtWidgets import QMessageBox, QDialog
from openpyxl import load_workbook
import requests
import time
import json

names, page_ids, page_url = [], [], []
total_count = 0
sleep = 1
class Checker:

    def __init__(self, params,wb,new):
       self.params = params
       self.Begin()
       if self.params["Excel"]:
            self.Create_table(wb,new)

       if self.params["JSON"]:
            self.Create_JSON(params["group_id"])

    def User_check(self,ids):
        users = ','.join(ids)
        u_url = f"https://api.vk.com/method/users.get?access_token={self.params['token']}&user_ids={users}&v=5.131"
        response = requests.get(u_url).json()

        for prof in response['response']:
            names.append(prof['first_name']+" "+prof['last_name'])
            page_ids.append(prof['id'])
            page_url.append("https://vk.com/"+str(prof['id']))

    def Check(self):
        global sleep

        url = f"https://api.vk.com/method/groups.getMembers?access_token={self.params['token']}&group_id={self.params['group_id']}&sort={self.params['sort']}&offset={self.params['offset']}&count=100&v=5.131"
        response = requests.get(url).json()
        count = response["response"]["count"]
        ids = []

        for obj in response["response"]["items"]:
            ids.append(str(obj))

        self.User_check(ids)

    def Begin(self):
        global total_count

        url = f"https://api.vk.com/method/groups.getMembers?access_token={self.params['token']}&group_id={self.params['group_id']}&sort={self.params['sort']}&offset={self.params['offset']}&count=100&v=5.131"
        response = requests.get(url).json()

        if "error" in response.keys():
            print(response["error"]["error_msg"])
            return

        total_count = response["response"]["count"]

        while total_count > self.params["offset"]:
            self.Check()
            self.params["offset"] += 100
            time.sleep(0.45)


    def Create_table(self,url,new):

        wb = load_workbook(url)

        if new == True:
            wb.remove(wb[wb.sheetnames[-1]])
            ws = wb.create_sheet("Members_Data")
        elif new == False:
            ws = wb[wb.sheetnames[-1]]

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        for i in range(1,len(names)+1):
            ws[f"A{i}"] = names[i-1]
            ws[f"B{i}"] = page_ids[i-1]
            ws[f"C{i}"] = page_url[i-1]

        wb.save(url)
        wb.close()

    def Create_JSON(self, url):
        base = {"info" : []}

        for i in range(len(names)):
            user = {"Name" : names[i], "ID" : page_ids[i], "Link" : page_url[i]}
            base["info"].append(user)

        with open(f"New_tables/{url}.json", "w", encoding="utf-8") as wf:
            wf.write(json.dumps(base, indent = 4, ensure_ascii=False))
